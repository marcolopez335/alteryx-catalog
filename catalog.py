"""
Alteryx Workflow Catalog Builder

Scans a directory of .yxmd/.yxwz/.yxmc files, parses the XML,
and builds a catalog with metadata, data lineage, and an auto-generated
plain-English description of what each workflow does.
"""

import argparse
import xml.etree.ElementTree as ET
from pathlib import Path

# Tool plugin names -> friendly category names
TOOL_CATEGORIES = {
    "DbFileInput": "Input",
    "DbFileOutput": "Output",
    "Join": "Join",
    "Filter": "Filter",
    "Formula": "Formula",
    "Summarize": "Summarize",
    "Sort": "Sort",
    "Select": "Select",
    "Union": "Union",
    "CrossTab": "Cross Tab",
    "Transpose": "Transpose",
    "RegEx": "RegEx",
    "DateTime": "DateTime",
    "MultiFieldFormula": "Multi-Field Formula",
    "Sample": "Sample",
    "Unique": "Unique",
    "Imputation": "Imputation",
    "GenerateRows": "Generate Rows",
    "RecordID": "Record ID",
    "MultiRowFormula": "Multi-Row Formula",
    "FindReplace": "Find Replace",
    "TextInput": "Text Input",
    "BrowseV2": "Browse",
    "AlteryxSelect": "Select",
    "Cleanse": "Data Cleansing",
    "Macro": "Macro",
    "RunCommand": "Run Command",
    "EmailOutput": "Email",
    "DownloadTool": "Download (API)",
    "BlockUntilDone": "Block Until Done",
    "Append": "Append Fields",
    "DynamicInput": "Dynamic Input",
    "DynamicReplace": "Dynamic Replace",
    "CharReplace": "Character Replace",
}


def _extract_tool_name(plugin: str) -> str:
    """Extract friendly tool name from plugin string."""
    parts = plugin.rsplit(".", 1)
    short = parts[-1] if parts else plugin
    return TOOL_CATEGORIES.get(short, short)


def _friendly_source(raw: str) -> str:
    """Turn a raw connection string / path into a short readable label."""
    if not raw:
        return "unknown"
    if "DSN=" in raw:
        for part in raw.split(";"):
            if part.strip().startswith("DSN="):
                return part.strip().split("=", 1)[1]
    if "\\" in raw or "/" in raw:
        return raw.rsplit("\\", 1)[-1].rsplit("/", 1)[-1]
    return raw


def _extract_connection_info(file_elem) -> str:
    """Pull connection string, file path, or table name from a File element.
    Shows full SQL — no truncation.
    """
    parts = []
    conn = file_elem.findtext("ConnectionString", "").strip()
    if conn:
        parts.append(conn)
    table = file_elem.findtext("Table", "").strip()
    if table:
        parts.append(table)
    if file_elem.text and file_elem.text.strip():
        parts.append(file_elem.text.strip())
    return " | ".join(parts) if parts else "unknown"


def _extract_sql(file_elem) -> str:
    """Extract just the SQL query from a File element, if present."""
    table = file_elem.findtext("Table", "").strip()
    if table and any(kw in table.upper() for kw in ("SELECT", "INSERT", "UPDATE", "DELETE", "WITH", "EXEC")):
        return table
    return ""


def parse_workflow(filepath: Path) -> dict:
    """Parse a single .yxmd file and extract catalog metadata."""
    tree = ET.parse(filepath)
    root = tree.getroot()

    info = {
        "file": str(filepath),
        "filename": filepath.name,
        "name": "",
        "description": "",
        "author": "",
        "created": "",
        "last_saved": "",
        "annotation": "",
        "inputs": [],
        "outputs": [],
        "tools": [],
        "tool_count": 0,
        "formulas": [],
        "filters": [],
        "joins": [],
        "summarizes": [],
        "sorts": [],
        "node_details": {},
        "connections": [],
    }

    # -- Metadata --
    meta = root.find(".//MetaInfo")
    if meta is not None:
        info["name"] = meta.findtext("Name", "").strip()
        info["description"] = meta.findtext("Description", "").strip()
        info["author"] = meta.findtext("Author", "").strip()
        info["created"] = meta.findtext("CreationDate", "").strip()
        info["last_saved"] = meta.findtext("LastSavedDate", "").strip()

    if not info["name"]:
        info["name"] = filepath.stem.replace("_", " ").title()

    # -- Workflow-level annotation --
    top_annot = root.find("./Properties/Annotation")
    if top_annot is not None:
        info["annotation"] = top_annot.findtext("DefaultAnnotationText", "").strip()

    # -- Connections (data flow graph) --
    for conn in root.findall(".//Connection"):
        origin = conn.find("Origin")
        dest = conn.find("Destination")
        if origin is not None and dest is not None:
            info["connections"].append({
                "from_id": origin.get("ToolID"),
                "to_id": dest.get("ToolID"),
            })

    # -- Nodes (tools) --
    nodes = root.findall(".//Node")
    info["tool_count"] = len(nodes)

    for node in nodes:
        gui = node.find("GuiSettings")
        if gui is None:
            continue

        tool_id = node.get("ToolID", "")
        plugin = gui.get("Plugin", "")
        tool_name = _extract_tool_name(plugin)
        if tool_name not in info["tools"]:
            info["tools"].append(tool_name)

        props = node.find("Properties")
        if props is None:
            continue

        config = props.find("Configuration")
        annotation = props.findtext(".//Annotation/Name", "").strip()
        if not annotation:
            annotation = props.findtext(".//Annotation/DefaultAnnotationText", "").strip()

        info["node_details"][tool_id] = {
            "plugin": plugin, "tool_name": tool_name, "annotation": annotation,
        }

        # -- Inputs --
        if "DbFileInput" in plugin or "TextInput" in plugin:
            input_info = {"tool_id": tool_id, "tool": tool_name, "annotation": annotation, "source": "", "sql": ""}
            if config is not None:
                file_elem = config.find("File")
                if file_elem is not None:
                    input_info["source"] = _extract_connection_info(file_elem)
                    input_info["source_short"] = _friendly_source(input_info["source"])
                    input_info["sql"] = _extract_sql(file_elem)
            info["inputs"].append(input_info)

        # -- Outputs --
        elif "DbFileOutput" in plugin or "EmailOutput" in plugin:
            output_info = {"tool_id": tool_id, "tool": tool_name, "annotation": annotation, "destination": ""}
            if config is not None:
                file_elem = config.find("File")
                if file_elem is not None:
                    output_info["destination"] = _extract_connection_info(file_elem)
                    output_info["dest_short"] = _friendly_source(output_info["destination"])
                    fmt = file_elem.get("FileFormat", "")
                    if "odbc" in output_info["destination"].lower() or fmt == "19":
                        output_info["type"] = "database"
                    elif output_info["destination"].lower().endswith(".csv"):
                        output_info["type"] = "CSV file"
                    elif output_info["destination"].lower().endswith((".xlsx", ".xls")):
                        output_info["type"] = "Excel file"
                    else:
                        output_info["type"] = "file"
            info["outputs"].append(output_info)

        # -- Filters --
        if "Filter" in plugin and config is not None:
            expr = config.findtext("Expression", "").strip()
            if expr:
                info["filters"].append({"tool_id": tool_id, "expression": expr, "annotation": annotation})

        # -- Joins --
        if "Join" in plugin and config is not None:
            join_fields = []
            for ji in config.findall(".//JoinInfo"):
                field = ji.find("Field")
                if field is not None:
                    join_fields.append(field.get("field", ""))
            if join_fields:
                info["joins"].append({
                    "tool_id": tool_id,
                    "fields": list(dict.fromkeys(join_fields)),
                    "annotation": annotation,
                })

        # -- Formulas --
        if "Formula" in plugin and config is not None:
            for ff in config.findall(".//FormulaField"):
                expr = ff.get("expression", "")
                field = ff.get("field", "")
                if expr:
                    info["formulas"].append({"field": field, "expression": expr})

        # -- Summarize --
        if "Summarize" in plugin and config is not None:
            for sf in config.findall(".//SummarizeField"):
                info["summarizes"].append({
                    "field": sf.get("field", ""),
                    "action": sf.get("action", ""),
                    "rename": sf.get("rename", ""),
                })

        # -- Sort --
        if "Sort" in plugin and config is not None:
            for sf in config.findall(".//SortInfo"):
                field = sf.get("field", "")
                order = sf.get("order", "Ascending")
                if field:
                    info["sorts"].append({"field": field, "order": order})

    return info


def auto_describe(wf: dict) -> str:
    """Generate a natural paragraph describing what the workflow does."""
    name = wf["name"]

    # --- Inputs ---
    input_labels = []
    for inp in wf["inputs"]:
        label = inp.get("annotation") or inp.get("source_short", "a data source")
        input_labels.append(label)

    # --- Build the opening sentence ---
    if len(input_labels) == 1:
        opening = f"This workflow pulls data from {input_labels[0]}"
    elif len(input_labels) == 2:
        opening = f"This workflow pulls data from {input_labels[0]} and {input_labels[1]}"
    elif input_labels:
        opening = f"This workflow pulls data from {', '.join(input_labels[:-1])}, and {input_labels[-1]}"
    else:
        opening = "This workflow processes data"

    # --- Transformation narrative ---
    transform_parts = []

    if wf["filters"]:
        for flt in wf["filters"]:
            if flt["annotation"]:
                transform_parts.append(f"filters to {flt['annotation'].lower()}")
            else:
                transform_parts.append(f"filters where {flt['expression']}")

    if wf["joins"]:
        for j in wf["joins"]:
            fields = ", ".join(j["fields"])
            if j["annotation"]:
                transform_parts.append(j["annotation"].lower())
            else:
                transform_parts.append(f"joins on {fields}")

    if wf["formulas"]:
        fields = [f["field"] for f in wf["formulas"] if f["field"]]
        if len(fields) <= 3:
            transform_parts.append(f"calculates {', '.join(fields)}")
        else:
            transform_parts.append(f"calculates {len(fields)} fields including {', '.join(fields[:3])}")

    if wf["summarizes"]:
        agg_parts = []
        for s in wf["summarizes"]:
            name_str = s["rename"] or s["field"]
            agg_parts.append(f"{s['action'].lower()} of {s['field']}")
        if len(agg_parts) <= 3:
            transform_parts.append(f"aggregates the {', '.join(agg_parts)}")
        else:
            transform_parts.append(f"aggregates {len(agg_parts)} measures")

    # Other notable tools
    for t in wf["tools"]:
        if t == "Data Cleansing":
            transform_parts.append("cleanses the data")
        elif t == "Cross Tab":
            transform_parts.append("pivots the data")
        elif t == "Transpose":
            transform_parts.append("transposes rows and columns")
        elif t == "Union":
            transform_parts.append("unions multiple data streams")
        elif t == "Unique":
            transform_parts.append("removes duplicate rows")
        elif t == "RegEx":
            transform_parts.append("applies regex transformations")
        elif t == "Download (API)":
            transform_parts.append("calls an external API")
        elif t == "Macro":
            transform_parts.append("calls a macro sub-workflow")

    # --- Output sentence ---
    output_parts = []
    for out in wf["outputs"]:
        dest = out.get("annotation") or out.get("dest_short", "a destination")
        out_type = out.get("type", "file")
        output_parts.append(f"{dest} ({out_type})")

    # --- Assemble the paragraph ---
    if transform_parts:
        middle = ", ".join(transform_parts)
        paragraph = f"{opening}, {middle}"
    else:
        paragraph = opening

    if len(output_parts) == 1:
        paragraph += f", and writes the results to {output_parts[0]}."
    elif len(output_parts) == 2:
        paragraph += f", and writes the results to {output_parts[0]} and {output_parts[1]}."
    elif output_parts:
        paragraph += f", and writes the results to {', '.join(output_parts[:-1])}, and {output_parts[-1]}."
    else:
        paragraph += "."

    # Email notification
    if "Email" in wf["tools"]:
        paragraph += " It also sends an email notification."

    # Run command
    if "Run Command" in wf["tools"]:
        paragraph += " It also runs an external system command."

    return paragraph


def build_summary(wf: dict) -> str:
    """Build a short flow-style summary: Input -> Transform -> Output."""
    parts = []
    if wf["inputs"]:
        sources = [inp.get("annotation") or inp.get("source_short", "?") for inp in wf["inputs"]]
        parts.append(f"Reads from: {', '.join(sources)}")
    transform_tools = [t for t in wf["tools"] if t not in ("Input", "Output", "Browse")]
    if transform_tools:
        parts.append(f"Transforms: {', '.join(transform_tools)}")
    if wf["outputs"]:
        dests = [out.get("annotation") or out.get("dest_short", "?") for out in wf["outputs"]]
        parts.append(f"Outputs to: {', '.join(dests)}")
    return " -> ".join(parts) if parts else "No summary available"


def print_catalog(workflows: list[dict]) -> None:
    """Print the catalog to the console using Rich."""
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich import box

    import sys, io
    if sys.platform == "win32":
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    console = Console()

    console.print()
    console.rule("[bold cyan]Alteryx Workflow Catalog[/]", style="cyan")
    console.print(f"  [dim]Scanned {len(workflows)} workflow(s)[/]")
    console.print()

    for wf in workflows:
        title_parts = [f"[bold white]{wf['name']}[/]"]
        if wf["author"]:
            title_parts.append(f"[dim]by {wf['author']}[/]")
        console.print(Panel("  ".join(title_parts), border_style="cyan", padding=(0, 2)))

        # Auto description — always shown
        console.print(f"  [bold]What it does:[/]")
        console.print(f"    {auto_describe(wf)}")
        console.print()

        if wf["description"]:
            console.print(f"  [bold]Author Note:[/]  {wf['description']}")
        if wf["annotation"]:
            console.print(f"  [bold]Department:[/]   {wf['annotation']}")
        console.print(f"  [bold]File:[/]         [dim]{wf['filename']}[/]")
        if wf["created"]:
            console.print(f"  [bold]Created:[/]      {wf['created']}")
        if wf["last_saved"]:
            console.print(f"  [bold]Last Saved:[/]   {wf['last_saved']}")
        console.print(f"  [bold]Tools:[/]        {wf['tool_count']} total ({', '.join(wf['tools'])})")

        # Inputs table
        if wf["inputs"]:
            console.print()
            t = Table(title="[bold green]Inputs[/]", box=box.SIMPLE, border_style="dim", show_lines=False)
            t.add_column("#", style="dim", width=3)
            t.add_column("Label", style="white", min_width=25)
            t.add_column("Source", style="cyan", no_wrap=False)
            for i, inp in enumerate(wf["inputs"], 1):
                t.add_row(str(i), inp["annotation"] or "-", inp["source"] or "-")
            console.print(t)

            # Show full SQL if any input has it
            for inp in wf["inputs"]:
                if inp.get("sql"):
                    label = inp["annotation"] or inp.get("source_short", "Input")
                    console.print(f"  [bold]SQL ({label}):[/]")
                    console.print(f"    [dim]{inp['sql']}[/]")
                    console.print()

        # Outputs table
        if wf["outputs"]:
            t = Table(title="[bold red]Outputs[/]", box=box.SIMPLE, border_style="dim", show_lines=False)
            t.add_column("#", style="dim", width=3)
            t.add_column("Label", style="white", min_width=25)
            t.add_column("Destination", style="yellow", no_wrap=False)
            t.add_column("Type", style="dim")
            for i, out in enumerate(wf["outputs"], 1):
                t.add_row(str(i), out["annotation"] or "-", out["destination"] or "-", out.get("type", "-"))
            console.print(t)

        # Filters
        if wf["filters"]:
            t = Table(title="[bold yellow]Filters[/]", box=box.SIMPLE, border_style="dim", show_lines=False)
            t.add_column("Label", style="white", min_width=20)
            t.add_column("Expression", style="dim")
            for f in wf["filters"]:
                t.add_row(f["annotation"] or "-", f["expression"])
            console.print(t)

        # Joins
        if wf["joins"]:
            t = Table(title="[bold blue]Joins[/]", box=box.SIMPLE, border_style="dim", show_lines=False)
            t.add_column("Label", style="white", min_width=20)
            t.add_column("Join Fields", style="cyan")
            for j in wf["joins"]:
                t.add_row(j["annotation"] or "-", ", ".join(j["fields"]))
            console.print(t)

        # Formulas
        if wf["formulas"]:
            t = Table(title="[bold magenta]Formulas[/]", box=box.SIMPLE, border_style="dim", show_lines=False)
            t.add_column("Field", style="white", min_width=15)
            t.add_column("Expression", style="dim")
            for f in wf["formulas"]:
                t.add_row(f["field"], f["expression"])
            console.print(t)

        # Summarize
        if wf["summarizes"]:
            t = Table(title="[bold cyan]Aggregations[/]", box=box.SIMPLE, border_style="dim", show_lines=False)
            t.add_column("Field", style="white", min_width=15)
            t.add_column("Action", style="yellow")
            t.add_column("Output Name", style="cyan")
            for s in wf["summarizes"]:
                t.add_row(s["field"], s["action"], s["rename"] or s["field"])
            console.print(t)

        console.print()
        console.rule(style="dim")
        console.print()


def export_excel(workflows: list[dict], output_path: Path) -> None:
    """Export catalog to a formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    # ===== Sheet 1: Catalog Overview =====
    ws = wb.active
    ws.title = "Catalog"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        bottom=Side(style="thin", color="D9E2F3"),
    )

    headers = [
        "Workflow Name", "Author", "Department", "What It Does",
        "File", "Created", "Last Saved", "Tool Count", "Tools Used",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, wf in enumerate(workflows, 2):
        values = [
            wf["name"],
            wf["author"],
            wf["annotation"],
            auto_describe(wf),
            wf["filename"],
            wf["created"],
            wf["last_saved"],
            wf["tool_count"],
            ", ".join(wf["tools"]),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = wrap
            cell.border = thin_border

    # Column widths
    widths = [30, 15, 25, 80, 30, 12, 12, 10, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # ===== Sheet 2: Inputs & Outputs =====
    ws2 = wb.create_sheet("Inputs & Outputs")
    headers2 = ["Workflow", "Direction", "Label", "Source / Destination", "Type", "SQL Query"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for wf in workflows:
        for inp in wf["inputs"]:
            values = [
                wf["name"], "Input",
                inp["annotation"] or "-",
                inp["source"],
                "-",
                inp.get("sql", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col, value=val)
                cell.alignment = wrap
                cell.border = thin_border
            row_idx += 1

        for out in wf["outputs"]:
            values = [
                wf["name"], "Output",
                out["annotation"] or "-",
                out["destination"],
                out.get("type", "-"),
                "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col, value=val)
                cell.alignment = wrap
                cell.border = thin_border
            row_idx += 1

    widths2 = [30, 10, 30, 60, 12, 80]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w
    ws2.freeze_panes = "A2"

    # ===== Sheet 3: Transformations =====
    ws3 = wb.create_sheet("Transformations")
    headers3 = ["Workflow", "Type", "Label / Field", "Detail"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for wf in workflows:
        for flt in wf["filters"]:
            for col, val in enumerate([wf["name"], "Filter", flt["annotation"] or "-", flt["expression"]], 1):
                cell = ws3.cell(row=row_idx, column=col, value=val)
                cell.alignment = wrap
                cell.border = thin_border
            row_idx += 1

        for j in wf["joins"]:
            for col, val in enumerate([wf["name"], "Join", j["annotation"] or "-", ", ".join(j["fields"])], 1):
                cell = ws3.cell(row=row_idx, column=col, value=val)
                cell.alignment = wrap
                cell.border = thin_border
            row_idx += 1

        for f in wf["formulas"]:
            for col, val in enumerate([wf["name"], "Formula", f["field"], f["expression"]], 1):
                cell = ws3.cell(row=row_idx, column=col, value=val)
                cell.alignment = wrap
                cell.border = thin_border
            row_idx += 1

        for s in wf["summarizes"]:
            detail = f"{s['action']}({s['field']})" + (f" as {s['rename']}" if s["rename"] else "")
            for col, val in enumerate([wf["name"], "Aggregation", s["rename"] or s["field"], detail], 1):
                cell = ws3.cell(row=row_idx, column=col, value=val)
                cell.alignment = wrap
                cell.border = thin_border
            row_idx += 1

    widths3 = [30, 12, 25, 60]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[ws3.cell(row=1, column=i).column_letter].width = w
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Exported catalog to {output_path}")


def export_csv(workflows: list[dict], output_path: Path) -> None:
    """Export catalog to a CSV file."""
    import csv

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Workflow Name", "File", "Author", "Department",
            "Description", "What It Does",
            "Created", "Last Saved", "Tool Count", "Tools Used",
            "Inputs", "Outputs", "Filters", "Joins", "Formulas",
            "Aggregations",
        ])
        for wf in workflows:
            inputs_str = "; ".join(
                f"{inp['annotation'] or '-'}: {inp['source']}" for inp in wf["inputs"]
            )
            outputs_str = "; ".join(
                f"{out['annotation'] or '-'}: {out['destination']}" for out in wf["outputs"]
            )
            filters_str = "; ".join(f["expression"] for f in wf["filters"])
            joins_str = "; ".join(
                f"{j['annotation'] or '-'}: {', '.join(j['fields'])}" for j in wf["joins"]
            )
            formulas_str = "; ".join(
                f"{f['field']} = {f['expression']}" for f in wf["formulas"]
            )
            agg_str = "; ".join(
                f"{s['action']}({s['field']})" for s in wf["summarizes"]
            )
            writer.writerow([
                wf["name"], wf["filename"], wf["author"], wf["annotation"],
                wf["description"], auto_describe(wf),
                wf["created"], wf["last_saved"],
                wf["tool_count"], ", ".join(wf["tools"]),
                inputs_str, outputs_str, filters_str, joins_str,
                formulas_str, agg_str,
            ])

    print(f"Exported catalog to {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Build a catalog of Alteryx workflows")
    parser.add_argument(
        "path", nargs="?", default=".",
        help="Directory to scan for .yxmd files (default: current directory)",
    )
    parser.add_argument(
        "--recursive", "-r", action="store_true",
        help="Scan subdirectories recursively",
    )
    parser.add_argument(
        "--csv", type=str, default=None,
        help="Export catalog to a CSV file",
    )
    parser.add_argument(
        "--excel", type=str, default=None,
        help="Export catalog to a formatted Excel file (.xlsx)",
    )
    args = parser.parse_args()

    scan_dir = Path(args.path)
    if not scan_dir.is_dir():
        print(f"Error: {scan_dir} is not a directory")
        return

    extensions = ("*.yxmd", "*.yxwz", "*.yxmc")
    files = []
    for ext in extensions:
        if args.recursive:
            files.extend(scan_dir.rglob(ext))
        else:
            files.extend(scan_dir.glob(ext))

    files = sorted(files)

    if not files:
        print(f"No Alteryx workflow files found in {scan_dir}")
        return

    workflows = []
    for f in files:
        try:
            wf = parse_workflow(f)
            workflows.append(wf)
        except ET.ParseError as e:
            print(f"  Warning: Failed to parse {f.name}: {e}")

    print_catalog(workflows)

    if args.csv:
        export_csv(workflows, Path(args.csv))

    if args.excel:
        export_excel(workflows, Path(args.excel))


if __name__ == "__main__":
    main()
